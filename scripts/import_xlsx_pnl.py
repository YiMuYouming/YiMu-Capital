#!/usr/bin/env python3
"""import_xlsx_pnl.py — 从 YiMu.xlsx 导入交易记录 → pnl_history.json

正向追踪法 + 今日锚定
"""
import json
from pathlib import Path
from datetime import datetime
from collections import defaultdict

XLSX_PATH = Path.home() / "Desktop" / "YiMu.xlsx"
ROOT_DIR = Path(__file__).resolve().parent.parent
PNL_PATH = ROOT_DIR / "data" / "pnl_history.json"
DASHBOARD_DATA = ROOT_DIR / "data" / "dashboard_data.json"
CURRENT_TOTAL_ASSET = 206075  # 当前总资产（来自 W16 报数）


def load_trades():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['交易记录']
    trades = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]: continue
        trades.append({
            'date': str(row[0]).strip()[:10],
            'code': str(row[2] or '').strip(),
            'name': str(row[3] or '').strip(),
            'action': str(row[4] or '').strip(),
            'qty': int(float(row[5] or 0)),
            'price': float(row[6] or 0),
            'amount': float(row[7] or 0),
            'trade_amount': float(row[8] or 0),
            'fee': float(row[9] or 0),
            'note': str(row[10] or ''),
        })
    return trades


def generate():
    trades = load_trades()
    by_date = defaultdict(list)
    for t in trades:
        by_date[t['date']].append(t)
    all_dates = sorted(by_date.keys())
    print(f"交易日: {all_dates[0]} ~ {all_dates[-1]} ({len(all_dates)} 天)")

    # 正向追踪现金 + 持仓成本
    cash_rel = 0
    positions = {}
    day_data = []

    for d in all_dates:
        realized_pnl = 0
        for t in by_date[d]:
            a = t['action']
            if a == '买入':
                cash_rel += t['amount']  # 负值
                c = t['code']
                if c:
                    old = positions.get(c, {})
                    oc = old.get('cost', 0) or 0
                    oq = old.get('qty', 0) or 0
                    positions[c] = {'cost': oc + abs(t['trade_amount']), 'qty': oq + t['qty']}
            elif a == '卖出':
                cash_rel += t['amount']  # 正值
                c = t['code']
                if c and c in positions:
                    old = positions[c]
                    r = t['qty'] / old['qty'] if old['qty'] > 0 else 0
                    cos = old['cost'] * r
                    realized_pnl += abs(t['trade_amount']) - cos - t['fee']
                    old['cost'] = old['cost'] - cos
                    old['qty'] = old['qty'] - t['qty']
            elif a == '速记':
                c = t['code']
                if c and t['trade_amount'] > 0:
                    positions[c] = {'cost': abs(t['trade_amount']), 'qty': t['qty']}

        pos_cost = sum(p['cost'] for p in positions.values() if p.get('qty', 0) > 0)
        day_data.append({
            'date': d,
            'cash_rel': round(cash_rel, 2),
            'pos_cost': round(pos_cost, 2),
            'realized_pnl': round(realized_pnl, 2),
        })

    # 锚定今日
    offset = CURRENT_TOTAL_ASSET - day_data[-1]['pos_cost'] - day_data[-1]['cash_rel']
    print(f"现金偏移: {offset:+,.2f}")

    # 正向重建 + TWR NAV
    records = []
    nav = 1.0

    for i, d in enumerate(day_data):
        cash_act = d['cash_rel'] + offset
        total = round(cash_act + d['pos_cost'], 2)

        if i == 0:
            pnl_pct = 0.0
            prev_total = total
        else:
            prev_total = records[i-1]['total']
            pnl_pct = round(d['realized_pnl'] / prev_total * 100, 4) if prev_total > 0 else 0

        if i > 0 and abs(pnl_pct) > 0.0001:
            nav = nav * (1 + pnl_pct / 100)

        records.append({
            'date': d['date'],
            'total': total,
            'nav': round(nav, 6),
            'pnl_pct': pnl_pct,
            'sh_pct': 0,
            'sz_pct': 0,
            'cy_pct': 0,
            'pos_pct': round(d['pos_cost'] / total * 100, 2) if total > 0 else 0,
            'deposit': 0,
        })

    start = day_data[0]
    start_total = round(start['cash_rel'] + offset + start['pos_cost'], 2)
    print(f"\n3/27 起始总资产: {start_total:,.0f}")
    print(f"今日总资产: {CURRENT_TOTAL_ASSET:,}")
    print(f"TWR 累计收益: {(nav-1)*100:+.2f}%")
    print(f"日均数: {len(records)}")

    # 写入 pnl_history.json
    history = {
        'meta': {
            'version': '1.0', 'currency': 'CNY',
            'total_deposit': round(start_total, 2),
            'last_twr_nav': nav,
            'updated': datetime.now().isoformat(),
        },
        'intraday_histories': {},
        'daily': records,
    }

    PNL_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(PNL_PATH, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)
    print(f"✅ pnl_history.json ({len(records)} 条)")

    # 同步 dashboard_data.json
    if DASHBOARD_DATA.exists():
        with open(DASHBOARD_DATA) as f:
            dd = json.load(f)
        dd['pnl'] = dd.get('pnl', {})
        dd['pnl']['总资产'] = CURRENT_TOTAL_ASSET
        dd['pnl']['累计入金'] = round(start_total, 2)
        with open(DASHBOARD_DATA, 'w', encoding='utf-8') as f:
            json.dump(dd, f, ensure_ascii=False, indent=2)
        print(f"✅ dashboard_data.json pnl 已更新")

    # 打印关键日
    print(f"\n{'日期':<12} {'总资产':>9} {'日收益%':>9} {'NAV':>9}")
    print("-" * 42)
    for r in records:
        if r['pnl_pct'] != 0 or r['date'] in [records[0]['date'], records[-1]['date']]:
            total_v = round(r['nav'] * start_total, 0)
            print(f"{r['date']:<12} {total_v:>9,.0f} {r['pnl_pct']:>+9.4f} {r['nav']:>9.6f}")


if __name__ == '__main__':
    generate()

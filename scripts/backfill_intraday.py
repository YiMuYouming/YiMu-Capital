#!/usr/bin/env python3
"""backfill_intraday.py — 历史日内5分钟K线回填 → SQLite intraday_snapshots

覆盖范围：2026-03-27 → 2026-05-15（37个交易日）
数据源：PyTDX get_security_bars(category=0, 5分钟K线)
"""
import json, sys, os, time
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import defaultdict

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from scripts.db import init_db
import sqlite3

XLSX_PATH = Path.home() / "Desktop" / "YiMu.xlsx"
CURRENT_TOTAL = 206075.0

# ===== 1. 解析 Excel 交易记录 =====

def load_trades():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['交易记录']
    trades = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]: continue
        trades.append({
            'date': str(row[0]).strip()[:10],
            'code': str(row[2] or '').strip(),
            'name': str(row[3] or '').strip(),
            'action': str(row[4] or '').strip(),
            'qty': int(float(row[5] or 0)),
            'price': float(row[6] or 0),
            'amount': float(row[7] or 0),
            'trade_amount': float(row[8] or 0),
            'fee': float(row[9] or 0),
        })
    return trades


# ===== 2. 构建每日持仓时间线 =====

def build_position_timeline(trades):
    """返回 {date: {code: {qty, cost, name}}}"""
    by_date = defaultdict(list)
    for t in trades:
        by_date[t['date']].append(t)

    all_dates = sorted(by_date.keys())
    positions = {}  # code → {qty, cost, name}
    timeline = {}

    for d in all_dates:
        for t in by_date[d]:
            a, c = t['action'], t['code']
            if a == '买入':
                if c:
                    old = positions.get(c, {'qty': 0, 'cost': 0.0, 'name': t['name']})
                    old['qty'] += t['qty']
                    old['cost'] += abs(t['trade_amount'])
                    positions[c] = old
            elif a == '卖出':
                if c and c in positions:
                    old = positions[c]
                    r = t['qty'] / old['qty'] if old['qty'] > 0 else 0
                    old['cost'] -= old['cost'] * r
                    old['qty'] -= t['qty']
            elif a == '速记':
                if c and t['trade_amount'] > 0:
                    positions[c] = {'qty': t['qty'], 'cost': abs(t['trade_amount']), 'name': t['name']}

        # 记录当天收盘持仓
        timeline[d] = {c: dict(p) for c, p in positions.items() if p['qty'] > 0}

    return timeline, all_dates


# ===== 3. 拉取 PyTDX 5分钟K线 =====

def get_tdx_api():
    from pytdx.hq import TdxHq_API
    servers = [
        ("110.41.147.114", 7709), ("119.147.212.81", 7709),
        ("124.70.176.52", 7709), ("47.100.236.28", 7709),
    ]
    for ip, port in servers:
        try:
            api = TdxHq_API()
            if api.connect(ip, port):
                return api
        except Exception:
            continue
    return None


def fetch_5min_bars(api, market, code, count=800, start=0, is_index=False):
    """拉取5分钟K线，返回 [(datetime_str, close_price), ...]"""
    try:
        if is_index:
            bars = api.get_index_bars(0, market, code, start, count)
        else:
            bars = api.get_security_bars(0, market, code, start, count)
        if not bars: return []
        result = []
        for b in bars:
            dt = f"{b['year']}-{b['month']:02d}-{b['day']:02d} {b['hour']:02d}:{b['minute']:02d}"
            result.append((dt, float(b['close'])))
        return result
    except Exception as e:
        print(f"  fetch error {code}: {e}")
        return []


def to_tdx_code(code):
    code = str(code).zfill(6)
    if code.startswith('6') or code.startswith('688'): return (1, code)
    elif code.startswith(('0','3')): return (0, code)
    return None


# ===== 4. 主流程 =====

def main():
    print("=" * 60)
    print("历史日内数据回填 (5分钟粒度)")
    print("=" * 60)

    # 4a. 构建持仓时间线
    trades = load_trades()
    timeline, trade_dates = build_position_timeline(trades)
    print(f"\n持仓时间线: {len(trade_dates)} 个交易日")

    # 4b. 计算总资产基线（锚定今日）
    # 正向追踪现金
    cash_rel = 0
    positions_cost = {}
    for d in trade_dates:
        for t in [t for t in trades if t['date'] == d]:
            if t['action'] == '买入': cash_rel += t['amount']
            elif t['action'] == '卖出': cash_rel += t['amount']
        pos_cost = sum(p['cost'] for p in timeline[d].values()) if d in timeline else 0

    last_pos_cost = sum(p['cost'] for p in timeline[trade_dates[-1]].values()) if timeline else 0
    offset = CURRENT_TOTAL - last_pos_cost - cash_rel
    print(f"现金偏移: {offset:+,.2f}")

    # 4c. 收集所有涉及过的股票
    all_stocks = set()
    for d in trade_dates:
        for code in timeline.get(d, {}):
            all_stocks.add(code)
    print(f"涉及股票: {len(all_stocks)} 只 → {sorted(all_stocks)}")

    # 4d. 连接 PyTDX 拉取 K 线
    api = get_tdx_api()
    if not api:
        print("❌ PyTDX 连接失败")
        return
    print("PyTDX 已连接")

    # 4e. 拉取所有股票的5分钟K线（分页）
    print("\n拉取股票5分钟K线...")
    stock_bars = {}  # code → {(date, time): close}
    for i, code in enumerate(sorted(all_stocks)):
        tdx = to_tdx_code(code)
        if not tdx: continue
        mkt, cd = tdx
        all_bars = []
        for start in [0, 800, 1600]:  # 翻页覆盖全部日期
            bars = fetch_5min_bars(api, mkt, cd, 800, start)
            if not bars: break
            all_bars.extend(bars)
        # 按 (date, time) 索引
        code_bars = {}
        for dt_str, close in all_bars:
            d, t = dt_str[:10], dt_str[11:16]
            code_bars[(d, t)] = close
        stock_bars[code] = code_bars
        print(f"  [{i+1}/{len(all_stocks)}] {code}: {len(code_bars)} 条5分钟K线")

    # 4f. 拉取三大指数5分钟K线
    print("\n拉取指数5分钟K线...")
    index_map = {'sh': (1, '000001'), 'sz': (0, '399001'), 'cy': (0, '399006')}
    index_close = {}  # name → {(date, time): pct_change}

    for name, (mkt, code) in index_map.items():
        all_bars = []
        for start in [0, 800, 1600]:
            bars = fetch_5min_bars(api, mkt, code, 800, start, is_index=True)
            if not bars: break
            all_bars.extend(bars)

        # 按日期分组，找每日第一根K线的close做基线
        day_bars = defaultdict(list)
        for dt_str, close in all_bars:
            d, t = dt_str[:10], dt_str[11:16]
            day_bars[d].append((t, close))

        idx_pct = {}
        for d, bars_list in day_bars.items():
            bars_list.sort()
            baseline = bars_list[0][1] if bars_list else 0
            for t, close in bars_list:
                pct = (close - baseline) / baseline * 100 if baseline > 0 else 0
                idx_pct[(d, t)] = round(pct, 4)

        index_close[name] = idx_pct
        print(f"  {name}: {len(idx_pct)} 条5分钟K线, {len(day_bars)} 天")

    api.disconnect()

    # 4g. 逐日逐5分钟计算 P&L
    print(f"\n计算P&L快照...")
    init_db()
    conn = sqlite3.connect(str(ROOT / "data" / "pnl.db"))
    cur = conn.cursor()

    total_snapshots = 0

    for d in trade_dates:
        pos = timeline.get(d, {})
        if not pos:
            continue

        total_cost = sum(p['cost'] for p in pos.values())
        total_asset = round(offset + cash_rel_at_date(d, trade_dates, trades, offset) + total_cost, 2)

        # 获取当天所有5分钟时间点（从任一支股票取时间轴）
        times = set()
        for code in pos:
            for (dd, tt) in stock_bars.get(code, {}):
                if dd == d:
                    times.add(tt)
        if not times:
            continue
        times = sorted(times)

        day_snapshots = 0
        for t in times:
            # 计算持仓市值
            mv = 0.0
            for code, p in pos.items():
                price = stock_bars.get(code, {}).get((d, t), 0)
                if price <= 0:
                    continue
                mv += p['qty'] * price
            if mv <= 0:
                continue

            pnl_pct = round((mv - total_cost) / total_asset * 100, 4) if total_asset > 0 else 0
            pos_pct = round(mv / total_asset * 100, 2) if total_asset > 0 else 0

            sh = index_close.get('sh', {}).get((d, t), 0)
            sz = index_close.get('sz', {}).get((d, t), 0)
            cy = index_close.get('cy', {}).get((d, t), 0)

            ts_iso = f"{d}T{t}:00"
            cur.execute("""
                INSERT OR REPLACE INTO intraday_snapshots
                (ts, date, pnl_pct, nav, sh_pct, sz_pct, cy_pct, pos_pct, mv, total_asset)
                VALUES (?, ?, ?, 1.0, ?, ?, ?, ?, ?, ?)
            """, (ts_iso, d, pnl_pct, sh, sz, cy, pos_pct, round(mv, 2), total_asset))
            day_snapshots += 1

        if day_snapshots > 0:
            conn.commit()
            total_snapshots += day_snapshots
            print(f"  {d}: {day_snapshots} 条快照 (持仓 {len(pos)} 只, 成本 {total_cost:,.0f})")

    conn.close()
    print(f"\n{'=' * 60}")
    print(f"✅ 回填完成: {total_snapshots} 条日内快照 ({trade_dates[0]} → {trade_dates[-1]})")


def cash_rel_at_date(target_date, trade_dates, trades, offset):
    """计算指定日期的相对现金"""
    cash = 0
    for d in trade_dates:
        if d > target_date: break
        for t in trades:
            if t['date'] == d:
                if t['action'] == '买入': cash += t['amount']
                elif t['action'] == '卖出': cash += t['amount']
    return cash


if __name__ == '__main__':
    main()

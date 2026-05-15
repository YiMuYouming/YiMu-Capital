#!/usr/bin/env python3
"""backfill_history.py — 从 YiMu.xlsx + PyTDX 历史K线重建 3/27→今天的完整日频数据

输出：daily_summary 表的完整历史记录（含上证/深证/创业指数）
"""
import json, sys, os
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import defaultdict

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = Path.home() / "Desktop" / "YiMu.xlsx"

# 导入 db
sys.path.insert(0, str(ROOT))
from scripts.db import init_db, insert_daily_summary


# ===== 1. 从 Excel 解析所有交易 =====

def load_trades():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['交易记录']
    trades = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]: continue
        trades.append({
            'date': str(row[0]).strip()[:10],
            'code': str(row[2] or '').strip(),
            'name': str(row[3] or '').strip(),
            'action': str(row[4] or '').strip(),
            'qty': int(float(row[5] or 0)),
            'price': float(row[6] or 0),
            'amount': float(row[7] or 0),
            'trade_amount': float(row[8] or 0),
            'fee': float(row[9] or 0),
        })
    return trades


# ===== 2. 正向追踪每日总资产 + TWR =====

def build_daily_assets(trades):
    """逐日追踪现金和持仓成本，计算每日总资产"""
    by_date = defaultdict(list)
    for t in trades:
        by_date[t['date']].append(t)

    all_dates = sorted(by_date.keys())
    print(f"交易日数: {len(all_dates)}")

    positions = {}  # code -> {cost, qty}
    cash_rel = 0    # 相对现金（从0开始的正向追踪）
    daily = []

    for d in all_dates:
        day_trades = by_date[d]
        realized_pnl = 0.0

        for t in day_trades:
            a = t['action']
            if a == '买入':
                cash_rel += t['amount']
                c = t['code']
                if c:
                    old = positions.get(c, {})
                    oc = old.get('cost', 0) or 0
                    oq = old.get('qty', 0) or 0
                    positions[c] = {'cost': oc + abs(t['trade_amount']), 'qty': oq + t['qty']}
            elif a == '卖出':
                cash_rel += t['amount']
                c = t['code']
                if c and c in positions:
                    old = positions[c]
                    r = t['qty'] / old['qty'] if old['qty'] > 0 else 0
                    cs = old['cost'] * r
                    realized_pnl += abs(t['trade_amount']) - cs - t['fee']
                    old['cost'] -= cs
                    old['qty'] -= t['qty']
            elif a == '速记':
                c = t['code']
                if c and t['trade_amount'] > 0:
                    positions[c] = {'cost': abs(t['trade_amount']), 'qty': t['qty']}

        pos_cost = sum(p['cost'] for p in positions.values() if p.get('qty', 0) > 0)
        daily.append({
            'date': d,
            'cash_rel': round(cash_rel, 2),
            'pos_cost': round(pos_cost, 2),
            'realized_pnl': round(realized_pnl, 2),
        })

    return daily


# ===== 3. 拉取历史指数数据 =====

def fetch_historical_index():
    """从 PyTDX 拉取 3/27→今天的上证/深证/创业日K线"""
    try:
        from pytdx.hq import TdxHq_API
    except ImportError:
        print("⚠️ PyTDX 未安装，指数数据将为空")
        return {}

    api = TdxHq_API()
    servers = [
        ("110.41.147.114", 7709), ("119.147.212.81", 7709),
        ("124.70.176.52", 7709), ("47.100.236.28", 7709),
    ]
    connected = False
    for ip, port in servers:
        try:
            if api.connect(ip, port):
                connected = True
                break
        except Exception:
            continue

    if not connected:
        print("⚠️ PyTDX 连接失败")
        return {}

    # 上证=1, 深证=0, 创业=0
    index_map = {'sh': (1, '000001'), 'sz': (0, '399001'), 'cy': (0, '399006')}
    result = {}

    for name, (mkt, code) in index_map.items():
        try:
            bars = api.get_index_bars(9, mkt, code, 0, 90)  # 日K, 近90天
            if bars:
                result[name] = {
                    str(b['datetime'])[:10] if isinstance(b['datetime'], int)
                    else (b.get('date', '') or b.get('datetime', ''))[:10]: {
                        'open': float(b['open']),
                        'close': float(b['close']),
                        'pct': round((float(b['close']) - float(b['open'])) / float(b['open']) * 100, 4)
                        if float(b['open']) > 0 else 0,
                    }
                    for b in bars
                }
                print(f"  {name}: {len(result[name])} 天K线")
        except Exception as e:
            print(f"  {name}: 获取失败 - {e}")

    api.disconnect()
    return result


# ===== 4. 整合 + 写入 =====

def main():
    trades = load_trades()
    daily = build_daily_assets(trades)

    # 锚定到当前总资产
    CURRENT_TOTAL = 206075.0
    last = daily[-1]
    offset = CURRENT_TOTAL - last['pos_cost'] - last['cash_rel']
    print(f"现金偏移: {offset:+,.2f}")

    # 拉取历史指数
    print("\n拉取历史指数数据...")
    idx_data = fetch_historical_index()

    # 计算 TWR + 写入
    init_db()
    prev_total = None
    nav = 1.0
    records = []

    for i, d in enumerate(daily):
        cash_act = d['cash_rel'] + offset
        total = round(cash_act + d['pos_cost'], 2)

        # 日收益率
        if i == 0 or not prev_total:
            pnl_pct = 0.0
        else:
            pnl_pct = round(d['realized_pnl'] / prev_total * 100, 4) if prev_total > 0 else 0

        if i > 0 and abs(pnl_pct) > 0.0001:
            nav = nav * (1 + pnl_pct / 100)

        # 获取当日指数
        sh = sz = cy = 0.0
        date_str = d['date']
        if idx_data:
            sh_d = idx_data.get('sh', {}).get(date_str, {})
            sz_d = idx_data.get('sz', {}).get(date_str, {})
            cy_d = idx_data.get('cy', {}).get(date_str, {})
            sh = sh_d.get('pct', 0)
            sz = sz_d.get('pct', 0)
            cy = cy_d.get('pct', 0)

        records.append({
            'date': date_str,
            'nav': round(nav, 6),
            'pnl_pct': pnl_pct,
            'sh_pct': sh,
            'sz_pct': sz,
            'cy_pct': cy,
            'pos_pct': round(d['pos_cost'] / total * 100, 1) if total > 0 else 0,
            'deposit': 0,
        })

        prev_total = total

        print(f"  {date_str}: asset={total:>10,.0f}  pnl={pnl_pct:>+8.4f}%  nav={nav:.6f}  sh={sh:+.2f}%")

    # 批量写入
    for r in records:
        insert_daily_summary(r)

    start_nav = records[0]['nav'] if records else 1.0
    print(f"\n{'='*50}")
    print(f"起始 NAV: {start_nav:.6f}")
    print(f"当前 NAV: {nav:.6f}")
    print(f"TWR 累计: {(nav-1)*100:+.2f}%")
    print(f"记录数: {len(records)} 条")
    print(f"✅ 历史数据已写入 daily_summary（含指数数据）")


if __name__ == '__main__':
    main()
